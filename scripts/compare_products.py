import os
import time
import re
import openpyxl
from openai import OpenAI
from datetime import datetime

EXCEL_FILE = os.getenv("EXCEL_PATH", "Batch 5 - 09 Jan-Vivek.xlsx")

def call_llm(prompt: str) -> str:
    """Send prompt to OpenAI model and return response text"""
    client = OpenAI(api_key=os.getenv("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"))
    model = os.getenv("MODEL_NAME", "gpt-4o-mini")

    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": "You are a precise comparison assistant. Follow the structure exactly."},
            {"role": "user", "content": prompt}
        ],
        temperature=0.2,
    )
    return response.choices[0].message.content.strip()

def parse_conclusion_and_final(text: str):
    """Extract Brand, Product, and Final result from AI output"""
    brand_match = re.search(r"Brand:\s*(Same|Different)", text, re.IGNORECASE)
    product_match = re.search(r"Product:\s*(Same|Different)", text, re.IGNORECASE)
    final_match = re.search(r"Final result:\s*(Same|Not same)", text, re.IGNORECASE)

    brand = brand_match.group(1).title() if brand_match else None
    product = product_match.group(1).title() if product_match else None
    final = final_match.group(1).title() if final_match else None

    return brand, product, final

def decide_status(brand: str, product: str) -> str:
    """Apply rule: Match only if both Brand and Product are Same"""
    if brand == "Same" and product == "Same":
        return "Match"
    return "Not Match"

def process_excel():
    """Process Excel file and update Status column using AI analysis"""
    try:
        print(f"Opening file: {EXCEL_FILE}")
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        print(f"Total rows: {ws.max_row}")

        # Add audit column for AI output if missing
        OUTPUT_COL = 15  # Column O for AI output
        if ws.cell(row=1, column=OUTPUT_COL).value is None:
            ws.cell(row=1, column=OUTPUT_COL).value = "AI_Output"

        for row in range(2, ws.max_row + 1):
            try:
                dell_product = ws.cell(row=row, column=4).value or ""
                snow_product = ws.cell(row=row, column=5).value or ""
                dell_publisher = ws.cell(row=row, column=6).value or ""
                snow_publisher = ws.cell(row=row, column=7).value or ""
                status_cell = ws.cell(row=row, column=11)
                prompt = ws.cell(row=row, column=14).value

                # Skip if already processed
                if status_cell.value in ["Match", "Not Match"]:
                    print(f"Row {row}: Already processed ({status_cell.value})")
                    continue

                # Skip if no prompt
                if not prompt:
                    print(f"Row {row}: No prompt, skipping")
                    continue

                print(f"\nRow {row}: Processing...")
                print(f"  Dell Product: {dell_product}")
                print(f"  SNOW Product: {snow_product}")
                print(f"  Dell Publisher: {dell_publisher}")
                print(f"  SNOW Publisher: {snow_publisher}")

                # Build full comparison prompt
                comparison_prompt = f"""
You are a comparison assistant. Compare two software/products/packages and return the result in the following fixed structure. 
**Do not add extra sections. Do not include code fences in the output. Do not use tables.**
**Write concise, clear bullets.**

Input pair:
- Item A: {dell_product} by {dell_publisher}
- Item B: {snow_product} by {snow_publisher}

**Required Output Structure (exactly in this order):**
1) Brand (Publisher) Check  
   - Clearly state the publisher for Item A and Item B, and whether the publishers are the same or different.  

2) Product Check  
   - Explain what Item A is and does.  
   - Explain what Item B is and does.  
   - Conclude whether the products are the same or different.  

3) Key Differences  
   - 3–6 bullet points capturing the most important differences.  

4) Summary  
   - 2–4 lines summarizing the relationship.  

5) Conclusion  
   - Two bullets only:
     - Brand: Same / Different
     - Product: Same / Different

6) Final result: Same / Not same
"""

                # Call AI
                llm_output = call_llm(comparison_prompt)
                brand, product, final = parse_conclusion_and_final(llm_output)
                status = decide_status(brand, product)

                # Write results
                status_cell.value = status
                ws.cell(row=row, column=OUTPUT_COL).value = llm_output

                print(f"  Status: {status} ✓")

                time.sleep(0.5)

            except Exception as e:
                print(f"Row {row}: Error - {str(e)}")
                ws.cell(row=row, column=11).value = "Error"
                ws.cell(row=row, column=OUTPUT_COL).value = f"ERROR: {e}"
                continue

        # Save the updated file
        wb.save(EXCEL_FILE)
        print(f"\n✅ SUCCESS! File updated: {EXCEL_FILE}")
        print(f"Total rows processed: {ws.max_row - 1}")

    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        raise

if __name__ == "__main__":
    process_excel()
