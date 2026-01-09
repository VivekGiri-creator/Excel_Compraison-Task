import openpyxl
import requests
import json
import os
import re

EXCEL_FILE = "Batch 5 - 08 Jan-Vivek.xlsx"
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")

# Comparison prompt template
COMPARISON_PROMPT = """You are a comparison assistant. Compare two software/products/packages. 

Input pair:  
- Item A: {item_a} by {publisher_a}
- Item B: {item_b} by {publisher_b}

**Required Output Structure:**

1) Brand (Publisher) Check  
   - State publisher for Item A and Item B
   - Same or different?

2) Product Check  
   - What is Item A?  (purpose, features, use)
   - What is Item B?  (purpose, features, use)
   - Same or different products?

3) Key Differences  
   - 3-6 bullet points of main differences

4) Summary  
   - 2-4 lines of relationship

5) Conclusion  
   - Brand:  Same / Different
   - Product: Same / Different

6) Final result: Same / Not same

**Rule:** If BOTH products are same AND publishers are same → Status: Match.  Otherwise → Not Match"""

def call_copilot_api(prompt):
    """Call GitHub Copilot API"""
    headers = {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Content-Type":  "application/json",
        "Accept": "application/vnd.github+json"
    }
    
    payload = {
        "model": "gpt-4",
        "messages": [
            {
                "role": "user",
                "content": prompt
            }
        ],
        "temperature": 0.7,
        "max_tokens": 1500
    }
    
    try:
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=30
        )
        
        if response.status_code == 200:
            result = response.json()
            return result. get("choices", [{}])[0].get("message", {}).get("content", "")
        else:
            print(f"API Error: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"Error:  {e}")
        return None

def get_final_status(comparison_text):
    """Extract final result and determine Match/Not Match"""
    if not comparison_text:
        return "Not Match"
    
    # Look for final result
    if "Final result:  Same" in comparison_text: 
        # Check if both product and brand are same
        if "Product: Same" in comparison_text and "Brand: Same" in comparison_text:
            return "Match"
    
    return "Not Match"

def process_excel():
    """Process Excel file"""
    try:
        print(f"Opening:  {EXCEL_FILE}")
        wb = openpyxl. load_workbook(EXCEL_FILE)
        ws = wb.active
        
        print(f"Total rows: {ws.max_row}")
        
        # Process each row (skip header row 1)
        for row in range(2, ws.max_row + 1):
            try:
                # Get values from columns
                dell_product = ws[f'D{row}'].value      # Column D
                snow_product = ws[f'E{row}'].value      # Column E
                dell_publisher = ws[f'F{row}'].value    # Column F
                snow_publisher = ws[f'G{row}']. value    # Column G
                status = ws[f'K{row}'].value            # Column K (Status)
                prompt = ws[f'L{row}'].value            # Column L (Prompt)
                
                # Skip if already done
                if status in ["Match", "Not Match"]:
                    print(f"Row {row}: Already processed ({status})")
                    continue
                
                if not prompt:
                    print(f"Row {row}: No prompt, skipping")
                    continue
                
                print(f"\nRow {row}: Processing...")
                print(f"  Product A: {dell_product}")
                print(f"  Product B: {snow_product}")
                
                # Create comparison prompt
                comp_prompt = COMPARISON_PROMPT.format(
                    item_a=dell_product or "Unknown",
                    publisher_a=dell_publisher or "Unknown",
                    item_b=snow_product or "Unknown",
                    publisher_b=snow_publisher or "Unknown"
                )
                
                # Call Copilot API
                print(f"  Calling Copilot API...")
                result = call_copilot_api(comp_prompt)
                
                if result:
                    status_value = get_final_status(result)
                    ws[f'K{row}'].value = status_value
                    print(f"  Status: {status_value} ✓")
                else:
                    ws[f'K{row}']. value = "Error"
                    print(f"  Status: Error")
                
            except Exception as e:
                print(f"Row {row}: Error - {e}")
                ws[f'K{row}'].value = "Error"
                continue
        
        # Save file
        wb.save(EXCEL_FILE)
        print(f"\n✅ Done!  File saved:  {EXCEL_FILE}")
        
    except Exception as e: 
        print(f"❌ Error:  {e}")

if __name__ == "__main__": 
    if not GITHUB_TOKEN:
        print("❌ GitHub token not found!")
        exit(1)
    
    process_excel()
